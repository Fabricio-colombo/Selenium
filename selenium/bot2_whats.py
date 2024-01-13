import openpyxl

# Carregando a planilha
planilha = openpyxl.load_workbook('bot_numeros.xlsx')

# Listando os nomes das folhas disponíveis
nomes_folhas = planilha.sheetnames
print("Nomes das folhas disponíveis:", nomes_folhas)

# Escolhendo a primeira folha (ou a folha que você deseja)
primeira_folha = planilha.active  # Ou substitua por planilha[nomes_folhas[0]] se souber o nome correto da folha

# Lendo a planilha
for linha in primeira_folha.iter_rows(values_only=True):
    # Seu código para processar cada linha
    nome = linha[0]
    numero = linha[1]
    # ...

# Fechar a planilha após a leitura
planilha.close()
